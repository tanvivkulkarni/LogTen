import argparse
import csv
import os
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Sequence, Tuple


@dataclass(frozen=True)
class TsvCell:
    row: int
    col: int
    text: str
    label: str
    x: int
    y: int
    x2: int
    y2: int


def _to_int(v: object, field: str) -> int:
    try:
        return int(str(v).strip())
    except Exception as e:
        raise ValueError(f"Expected int for {field}, got {v!r}") from e


def read_tsv_cells(tsv_path: Path) -> List[TsvCell]:
    with open(tsv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        rows = list(reader)
    if not rows:
        return []

    required = ["row", "column", "label", "x", "y", "x2", "y2", "text"]
    missing = [k for k in required if k not in rows[0]]
    if missing:
        raise ValueError(f"{tsv_path}: missing columns {missing}; found {list(rows[0].keys())}")

    out: List[TsvCell] = []
    for r in rows:
        out.append(
            TsvCell(
                row=_to_int(r.get("row"), "row"),
                col=_to_int(r.get("column"), "column"),
                label=str(r.get("label") or ""),
                x=_to_int(r.get("x"), "x"),
                y=_to_int(r.get("y"), "y"),
                x2=_to_int(r.get("x2"), "x2"),
                y2=_to_int(r.get("y2"), "y2"),
                text=str(r.get("text") or ""),
            )
        )
    return out


def _is_blank(s: str) -> bool:
    return not s or not str(s).strip()


_TIME_RE = re.compile(r"^\s*\d{1,2}:\d{2}\s*$")
_ZERO_TIME_RE = re.compile(r"^\s*0{1,2}:\d{2}\s*$")
_NUMERICISH_RE = re.compile(r"^\s*[\d\.\-/:]+\s*$")


def _cell_is_headerish(text: str) -> bool:
    """Heuristic: header cells often have letters."""
    t = (text or "").strip()
    if not t:
        return False
    # Route, Aircraft, PIC, Instrument, etc.
    return any(ch.isalpha() for ch in t)


def _cell_is_dataish(text: str) -> bool:
    """Heuristic: data rows often have times/zeros/numbers."""
    t = (text or "").strip()
    if not t:
        return False
    if _TIME_RE.match(t) or _ZERO_TIME_RE.match(t):
        return True
    if _NUMERICISH_RE.match(t) and not any(ch.isalpha() for ch in t):
        return True
    return False


def build_grid(cells: Sequence[TsvCell], join_duplicates_with: str = " ") -> Dict[int, Dict[int, str]]:
    grid: Dict[int, Dict[int, str]] = defaultdict(dict)
    for c in cells:
        v = c.text or ""
        if c.col in grid[c.row]:
            existing = grid[c.row][c.col]
            if join_duplicates_with is None:
                grid[c.row][c.col] = v
            else:
                if existing.strip() and v.strip():
                    grid[c.row][c.col] = existing + join_duplicates_with + v
                else:
                    grid[c.row][c.col] = existing + v
        else:
            grid[c.row][c.col] = v
    return grid


def _row_non_empty_count(row_map: Dict[int, str]) -> int:
    return sum(1 for v in row_map.values() if not _is_blank(v))


def _row_max_col(row_map: Dict[int, str]) -> int:
    return max(row_map.keys(), default=0)


def pick_main_table_rows(
    grid: Dict[int, Dict[int, str]],
    *,
    min_fraction_of_max_cols: float = 0.60,
    min_non_empty: int = 6,
) -> List[int]:
    """Pick a contiguous block of rows that looks like the main table."""
    if not grid:
        return []

    rows_sorted = sorted(grid.keys())
    max_cols = max((_row_max_col(grid[r]) for r in rows_sorted), default=0)
    if max_cols <= 0:
        return []

    threshold_cols = max(1, int(max_cols * min_fraction_of_max_cols))

    candidate_rows: List[int] = []
    for r in rows_sorted:
        row_map = grid[r]
        if _row_max_col(row_map) >= threshold_cols and _row_non_empty_count(row_map) >= min_non_empty:
            candidate_rows.append(r)

    if not candidate_rows:
        return []

    # Choose the longest contiguous segment
    best: List[int] = []
    cur: List[int] = [candidate_rows[0]]
    for rr in candidate_rows[1:]:
        if rr == cur[-1] + 1:
            cur.append(rr)
        else:
            if len(cur) > len(best):
                best = cur
            cur = [rr]
    if len(cur) > len(best):
        best = cur
    return best


def _header_score(values: Sequence[str]) -> int:
    headerish = sum(1 for v in values if _cell_is_headerish(v))
    dataish = sum(1 for v in values if _cell_is_dataish(v))
    return headerish - dataish


def pick_header_row_near_table(grid: Dict[int, Dict[int, str]], table_rows: Sequence[int]) -> Optional[int]:
    """
    Pick the most header-like row near the table block.

    Strategy:
    - Prefer rows *above* the first detected data row (often where the column names live).
    - If none look good, fall back to the best row *within* the table block.
    """
    if not table_rows:
        return None

    first = table_rows[0]
    max_col_table = max((_row_max_col(grid.get(r, {})) for r in table_rows), default=0)
    min_cols_for_header = max(1, int(max_col_table * 0.40))

    # Search a window above the table for headers (up to 15 rows back).
    start_search = max(1, first - 15)
    search_rows = list(range(start_search, first))

    best_row = None
    best_score = None

    def consider(r: int) -> None:
        nonlocal best_row, best_score
        row_map = grid.get(r, {})
        if _row_max_col(row_map) < min_cols_for_header:
            return
        values = [v for _, v in sorted(row_map.items()) if not _is_blank(v)]
        if len(values) < 4:
            return
        score = _header_score(values)
        if best_score is None or score > best_score:
            best_score = score
            best_row = r

    for r in search_rows:
        consider(r)

    if best_row is not None:
        return best_row

    # Fallback: best inside table block
    for r in table_rows:
        consider(r)

    return best_row or table_rows[0]


def _sanitize_col_name(name: str) -> str:
    n = " ".join((name or "").split()).strip()
    if not n:
        return ""
    # keep it excel-friendly
    n = n.replace("\n", " ")
    n = re.sub(r"[^\w\s\-/().]", "", n)
    n = re.sub(r"\s+", " ", n).strip()
    return n


def make_column_names(grid: Dict[int, Dict[int, str]], header_row: int, cols: Sequence[int]) -> List[str]:
    seen: Dict[str, int] = {}
    out: List[str] = []
    header_map = grid.get(header_row, {})
    for c in cols:
        raw = header_map.get(c, "")
        name = _sanitize_col_name(raw)
        if not name:
            name = f"C{c}"
        base = name
        if base in seen:
            seen[base] += 1
            name = f"{base}_{seen[base]}"
        else:
            seen[base] = 1
        out.append(name)
    return out


def write_nice_csv_from_tsv(
    tsv_path: Path,
    out_path: Path,
    *,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
    data_end_row: Optional[int] = None,
    join_duplicates_with: str = " ",
    include_row_number_column: bool = False,
) -> Tuple[Optional[int], List[int]]:
    cells = read_tsv_cells(tsv_path)
    grid = build_grid(cells, join_duplicates_with=join_duplicates_with)

    table_rows = pick_main_table_rows(grid)
    if not table_rows:
        raise RuntimeError(f"{tsv_path}: could not detect a main table block (try manual --header-row/--data-start-row)")

    auto_header = pick_header_row_near_table(grid, table_rows)
    hdr = header_row if header_row is not None else auto_header
    if hdr is None:
        hdr = table_rows[0]

    start = data_start_row if data_start_row is not None else (hdr + 1)
    end = data_end_row if data_end_row is not None else table_rows[-1]
    if end < start:
        end = start

    # Columns: use the max col observed in the table block (keeps alignment)
    max_col = 0
    for r in table_rows:
        max_col = max(max_col, _row_max_col(grid.get(r, {})))
    cols = list(range(1, max_col + 1))
    col_names = make_column_names(grid, hdr, cols)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with open(out_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        header = (["row"] if include_row_number_column else []) + col_names
        writer.writerow(header)
        for r in range(start, end + 1):
            row_map = grid.get(r, {})
            row_vals = [row_map.get(c, "") for c in cols]
            if all(_is_blank(v) for v in row_vals):
                continue
            writer.writerow(([r] if include_row_number_column else []) + row_vals)

    return hdr, table_rows


def write_nice_xlsx_from_tsv(
    tsv_path: Path,
    out_path: Path,
    *,
    header_row: Optional[int] = None,
    data_start_row: Optional[int] = None,
    data_end_row: Optional[int] = None,
    join_duplicates_with: str = " ",
    include_row_number_column: bool = False,
    sheet_name: str = "table",
) -> Tuple[Optional[int], List[int]]:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError("openpyxl is required. Install with: pip install openpyxl") from e

    cells = read_tsv_cells(tsv_path)
    grid = build_grid(cells, join_duplicates_with=join_duplicates_with)

    table_rows = pick_main_table_rows(grid)
    if not table_rows:
        raise RuntimeError(f"{tsv_path}: could not detect a main table block (try manual --header-row/--data-start-row)")

    auto_header = pick_header_row_near_table(grid, table_rows)
    hdr = header_row if header_row is not None else auto_header
    if hdr is None:
        hdr = table_rows[0]

    start = data_start_row if data_start_row is not None else (hdr + 1)
    end = data_end_row if data_end_row is not None else table_rows[-1]
    if end < start:
        end = start

    max_col = 0
    for r in table_rows:
        max_col = max(max_col, _row_max_col(grid.get(r, {})))
    cols = list(range(1, max_col + 1))
    col_names = make_column_names(grid, hdr, cols)

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # header row
    col_offset = 2 if include_row_number_column else 1
    if include_row_number_column:
        ws.cell(row=1, column=1, value="row")
    for j, name in enumerate(col_names, start=col_offset):
        ws.cell(row=1, column=j, value=name)

    out_r = 2
    for r in range(start, end + 1):
        row_map = grid.get(r, {})
        row_vals = [row_map.get(c, "") for c in cols]
        if all(_is_blank(v) for v in row_vals):
            continue
        if include_row_number_column:
            ws.cell(row=out_r, column=1, value=r)
        for j, v in enumerate(row_vals, start=col_offset):
            if v != "":
                ws.cell(row=out_r, column=j, value=v)
        out_r += 1

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return hdr, table_rows


def _collect_tsvs(path: Path) -> List[Path]:
    path = path.resolve()
    if path.is_file():
        return [path] if path.suffix.lower() == ".tsv" else []
    if not path.is_dir():
        return []
    return sorted([p for p in path.iterdir() if p.is_file() and p.suffix.lower() == ".tsv"])


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate a clean (header + data) CSV/XLSX from new_sample_ocr.py TSV output."
    )
    parser.add_argument(
        "--input",
        "-i",
        required=True,
        help="Path to a .tsv file OR a directory containing .tsv files (e.g. output_today/df)",
    )
    parser.add_argument(
        "--out-dir",
        default=None,
        help="Output directory (default: alongside input TSVs)",
    )
    parser.add_argument(
        "--format",
        choices=["csv", "xlsx", "both"],
        default="both",
        help="Output format (default: both)",
    )
    parser.add_argument(
        "--header-row",
        type=int,
        default=None,
        help="Manual header row number (1-based, matches TSV 'row')",
    )
    parser.add_argument(
        "--data-start-row",
        type=int,
        default=None,
        help="Manual data start row number (1-based)",
    )
    parser.add_argument(
        "--data-end-row",
        type=int,
        default=None,
        help="Manual data end row number (1-based)",
    )
    parser.add_argument(
        "--include-row-number",
        action="store_true",
        help="Include the TSV row number as first column in output",
    )
    parser.add_argument(
        "--join-duplicates-with",
        default=" ",
        help="If multiple boxes share the same (row,column), join their text with this separator",
    )

    args = parser.parse_args()
    in_path = Path(args.input)
    tsvs = _collect_tsvs(in_path)
    if not tsvs:
        raise SystemExit(f"No .tsv files found at: {in_path}")

    out_dir = Path(args.out_dir).resolve() if args.out_dir else None

    for tsv in tsvs:
        base_out_dir = out_dir if out_dir else tsv.parent
        stem = tsv.stem
        csv_out = base_out_dir / f"{stem}_nice.csv"
        xlsx_out = base_out_dir / f"{stem}_nice.xlsx"

        hdr = None
        table_rows: List[int] = []

        if args.format in ("csv", "both"):
            hdr, table_rows = write_nice_csv_from_tsv(
                tsv,
                csv_out,
                header_row=args.header_row,
                data_start_row=args.data_start_row,
                data_end_row=args.data_end_row,
                join_duplicates_with=args.join_duplicates_with,
                include_row_number_column=args.include_row_number,
            )
            print(f"Wrote: {csv_out}")

        if args.format in ("xlsx", "both"):
            hdr, table_rows = write_nice_xlsx_from_tsv(
                tsv,
                xlsx_out,
                header_row=args.header_row,
                data_start_row=args.data_start_row,
                data_end_row=args.data_end_row,
                join_duplicates_with=args.join_duplicates_with,
                include_row_number_column=args.include_row_number,
            )
            print(f"Wrote: {xlsx_out}")

        if hdr is not None and table_rows:
            print(f"{tsv.name}: detected table rows {table_rows[0]}..{table_rows[-1]}, header_row={hdr}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())

