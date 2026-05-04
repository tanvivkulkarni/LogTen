import argparse
import csv
import os
from collections import defaultdict
from typing import Dict, List


def _read_tsv_rows(tsv_path: str) -> List[Dict[str, str]]:
    with open(tsv_path, "r", newline="", encoding="utf-8") as f:
        reader = csv.DictReader(f, delimiter="\t")
        rows = list(reader)
    return rows


def write_flat_csv(tsv_path: str, csv_path: str) -> None:
    rows = _read_tsv_rows(tsv_path)
    if not rows:
        raise ValueError(f"No rows found in TSV: {tsv_path}")

    fieldnames = list(rows[0].keys())
    os.makedirs(os.path.dirname(os.path.abspath(csv_path)) or ".", exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_text_boxes_csv(
    tsv_path: str,
    csv_path: str,
    *,
    keep_empty_text: bool = True,
    row_field: str = "row",
    col_field: str = "column",
    label_field: str = "label",
    x_field: str = "x",
    y_field: str = "y",
    x2_field: str = "x2",
    y2_field: str = "y2",
    text_field: str = "text",
) -> None:
    rows = _read_tsv_rows(tsv_path)
    if not rows:
        raise ValueError(f"No rows found in TSV: {tsv_path}")

    fieldnames = [
        row_field,
        col_field,
        label_field,
        x_field,
        y_field,
        x2_field,
        y2_field,
        text_field,
    ]

    missing = [f for f in fieldnames if f not in rows[0]]
    if missing:
        raise ValueError(
            f"Missing required TSV columns: {missing}. Found: {list(rows[0].keys())}"
        )

    out_rows: List[Dict[str, str]] = []
    for r in rows:
        text_val = r.get(text_field, "")
        if text_val is None:
            text_val = ""
        if not keep_empty_text and str(text_val).strip() == "":
            continue

        out_rows.append({k: ("" if r.get(k) is None else str(r.get(k))) for k in fieldnames})

    os.makedirs(os.path.dirname(os.path.abspath(csv_path)) or ".", exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(out_rows)


def _to_int(value: str, field: str) -> int:
    try:
        return int(str(value).strip())
    except Exception as e:
        raise ValueError(f"Expected integer for '{field}', got {value!r}") from e


def write_grid_csv(
    tsv_path: str,
    csv_path: str,
    *,
    row_field: str = "row",
    col_field: str = "column",
    value_field: str = "text",
    include_missing_columns: bool = True,
    join_duplicates_with: str = " ",
) -> None:
    rows = _read_tsv_rows(tsv_path)
    if not rows:
        raise ValueError(f"No rows found in TSV: {tsv_path}")

    grid: Dict[int, Dict[int, str]] = defaultdict(dict)
    row_values = set()
    col_values = set()

    for r in rows:
        rr = _to_int(r.get(row_field, ""), row_field)
        cc = _to_int(r.get(col_field, ""), col_field)
        val = r.get(value_field, "")
        if val is None:
            val = ""
        s = str(val)
        if cc in grid[rr] and join_duplicates_with is not None:
            existing = grid[rr][cc]
            if existing.strip() and s.strip():
                grid[rr][cc] = existing + join_duplicates_with + s
            else:
                grid[rr][cc] = existing + s
        else:
            grid[rr][cc] = s
        row_values.add(rr)
        col_values.add(cc)

    sorted_rows = sorted(row_values)
    sorted_cols = sorted(col_values)
    if include_missing_columns and sorted_cols:
        min_c, max_c = sorted_cols[0], sorted_cols[-1]
        sorted_cols = list(range(min_c, max_c + 1))

    header = ["row"] + [f"C{c}" for c in sorted_cols]
    os.makedirs(os.path.dirname(os.path.abspath(csv_path)) or ".", exist_ok=True)
    with open(csv_path, "w", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        writer.writerow(header)
        for rr in sorted_rows:
            out_row = [rr]
            row_map = grid.get(rr, {})
            for cc in sorted_cols:
                out_row.append(row_map.get(cc, ""))
            writer.writerow(out_row)


def write_grid_xlsx(
    tsv_path: str,
    xlsx_path: str,
    *,
    row_field: str = "row",
    col_field: str = "column",
    value_field: str = "text",
    include_missing_columns: bool = True,
    join_duplicates_with: str = " ",
    sheet_name: str = "grid",
) -> None:
    try:
        from openpyxl import Workbook
    except Exception as e:
        raise RuntimeError(
            "openpyxl is required to write .xlsx. Install with: pip install openpyxl"
        ) from e

    rows = _read_tsv_rows(tsv_path)
    if not rows:
        raise ValueError(f"No rows found in TSV: {tsv_path}")

    grid: Dict[int, Dict[int, str]] = defaultdict(dict)
    row_values = set()
    col_values = set()

    for r in rows:
        rr = _to_int(r.get(row_field, ""), row_field)
        cc = _to_int(r.get(col_field, ""), col_field)
        val = r.get(value_field, "")
        if val is None:
            val = ""
        s = str(val)
        if cc in grid[rr] and join_duplicates_with is not None:
            existing = grid[rr][cc]
            if existing.strip() and s.strip():
                grid[rr][cc] = existing + join_duplicates_with + s
            else:
                grid[rr][cc] = existing + s
        else:
            grid[rr][cc] = s
        row_values.add(rr)
        col_values.add(cc)

    sorted_rows = sorted(row_values)
    sorted_cols = sorted(col_values)
    if include_missing_columns and sorted_cols:
        min_c, max_c = sorted_cols[0], sorted_cols[-1]
        sorted_cols = list(range(min_c, max_c + 1))

    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name

    # Header row
    ws.cell(row=1, column=1, value="row")
    for j, cc in enumerate(sorted_cols, start=2):
        ws.cell(row=1, column=j, value=f"C{cc}")

    # Data
    for i, rr in enumerate(sorted_rows, start=2):
        ws.cell(row=i, column=1, value=rr)
        row_map = grid.get(rr, {})
        for j, cc in enumerate(sorted_cols, start=2):
            v = row_map.get(cc, "")
            if v != "":
                ws.cell(row=i, column=j, value=v)

    os.makedirs(os.path.dirname(os.path.abspath(xlsx_path)) or ".", exist_ok=True)
    wb.save(xlsx_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Convert TSV (row/column/x/y/x2/y2/text) to CSV."
    )
    parser.add_argument("--input", "-i", required=True, help="Path to input .tsv file")
    parser.add_argument(
        "--output",
        "-o",
        help="Path to output .csv file. If omitted, uses <input>.csv",
    )
    parser.add_argument(
        "--mode",
        choices=["flat", "grid", "grid_xlsx", "text_boxes"],
        default="flat",
        help="flat = same rows/columns; grid = pivot text into row/column table; grid_xlsx = same but .xlsx; text_boxes = only row/column/label/coords/text",
    )
    parser.add_argument(
        "--row-field",
        default="row",
        help="TSV column name for row index (default: row)",
    )
    parser.add_argument(
        "--col-field",
        default="column",
        help="TSV column name for column index (default: column)",
    )
    parser.add_argument(
        "--value-field",
        default="text",
        help="TSV column name to place in grid cells (default: text)",
    )
    parser.add_argument(
        "--no-fill-missing-columns",
        action="store_true",
        help="(grid/grid_xlsx) do not fill missing columns between min and max",
    )
    parser.add_argument(
        "--join-duplicates-with",
        default=" ",
        help="(grid/grid_xlsx) if multiple rows share same (row,column), join texts with this separator",
    )
    parser.add_argument(
        "--drop-empty-text",
        action="store_true",
        help="(text_boxes mode) skip rows where text is empty/whitespace",
    )
    parser.add_argument(
        "--label-field",
        default="label",
        help="(text_boxes mode) TSV column name for label (default: label)",
    )
    parser.add_argument(
        "--x-field",
        default="x",
        help="(text_boxes mode) TSV column name for left x (default: x)",
    )
    parser.add_argument(
        "--y-field",
        default="y",
        help="(text_boxes mode) TSV column name for top y (default: y)",
    )
    parser.add_argument(
        "--x2-field",
        default="x2",
        help="(text_boxes mode) TSV column name for right x (default: x2)",
    )
    parser.add_argument(
        "--y2-field",
        default="y2",
        help="(text_boxes mode) TSV column name for bottom y (default: y2)",
    )

    args = parser.parse_args()
    in_path = args.input
    out_path = args.output or os.path.splitext(in_path)[0] + ".csv"

    if args.mode == "flat":
        write_flat_csv(in_path, out_path)
    elif args.mode == "grid":
        write_grid_csv(
            in_path,
            out_path,
            row_field=args.row_field,
            col_field=args.col_field,
            value_field=args.value_field,
            include_missing_columns=not args.no_fill_missing_columns,
            join_duplicates_with=args.join_duplicates_with,
        )
    elif args.mode == "grid_xlsx":
        write_grid_xlsx(
            in_path,
            out_path,
            row_field=args.row_field,
            col_field=args.col_field,
            value_field=args.value_field,
            include_missing_columns=not args.no_fill_missing_columns,
            join_duplicates_with=args.join_duplicates_with,
        )
    else:
        write_text_boxes_csv(
            in_path,
            out_path,
            keep_empty_text=not args.drop_empty_text,
            row_field=args.row_field,
            col_field=args.col_field,
            label_field=args.label_field,
            x_field=args.x_field,
            y_field=args.y_field,
            x2_field=args.x2_field,
            y2_field=args.y2_field,
            text_field=args.value_field,
        )

    print(f"Wrote: {out_path}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

